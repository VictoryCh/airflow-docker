import datetime
import logging
import time

import openpyxl
import xlwt
from airflow.models import BaseOperator
from airflow.models import Variable
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.utils.decorators import apply_defaults
from cryptography.fernet import Fernet
from lxml import etree as lxml


# import xml.etree.ElementTree as ET


class Postgres2SMB_decrypt(BaseOperator):
    template_fields = ('sql',)
    template_ext = ('.sql',)  # file format
    ui_color = '#e08c8c'
    dest_cols = []

    @apply_defaults
    def __init__(self, sql, postgres_conn_id=None, parameters=None, autocommit=False, rows_chunk=5000,
                 samba_conn_id=None, headings=None, share=None, dir_samba=None, file_name='test', file_format='csv', delimiter=';',
                 encoding='utf-8', typ=None, modewr='wb', SheetName='Лист1', sheet_name_sql=None, template_excel_xml=None,
                 decrypt_col=None, *args, **kwargs):
        super(Postgres2SMB_decrypt, self).__init__(*args, **kwargs)
        if parameters is None:
            parameters = {}
        self.sql = sql  # sql query on source system (here, if necessary, transformations are performed)
        self.postgres_conn_id = postgres_conn_id
        self.parameters = parameters
        self.autocommit = autocommit
        self.rows_chunk = rows_chunk
        self.samba_conn_id = samba_conn_id
        self.headings = headings
        self.share = share
        self.dir_samba = dir_samba
        self.file_name = file_name
        self.file_format = file_format
        self.delimiter = delimiter
        self.encoding = encoding
        self.typ = typ
        self.modewr = modewr
        self.SheetName = SheetName
        self.sheet_name_sql = sheet_name_sql
        self.template_excel_xml = template_excel_xml
        self.decrypt_col = decrypt_col # пишем из основного запроса
        self.outputs = {'xml': self.give_excel_xml_output,  # self.give_xml_output,
                        'xls': self.give_xls_output,
                        'xlsx': self.give_xlsx_output
                        }

    def _execute(self, src_hook, dest_hook, context):
        with src_hook.get_conn() as src_conn:
            logging.info("Source id connection: %s", self.postgres_conn_id)
            logging.info("Executing sql = %s", self.sql)
            dirr = '' if not self.dir_samba else self.dir_samba + '/'
            dir_file = f'{dirr}{self.file_name}.{self.file_format}'
            q_rows = self.outputs[self.file_format](src_conn=src_conn, dir_file=dir_file, context=context)
            # q_rows = self.give_xls_output(src_conn, dir_file)

            logging.info("Finished data transfer")
            context['task_instance'].xcom_push(key='q_rows', value=q_rows)

    def execute(self, context):
        # time.sleep(5)
        dest_hook = 'SambaHook(samba_conn_id=self.samba_conn_id, share=self.share)'
        src_hook = PostgresHook(postgres_conn_id=self.postgres_conn_id)
        try:
            rows_count = context['task_instance'].xcom_pull(key='full_qr')
            full_comment = context['task_instance'].xcom_pull(key='full_comment')
            if rows_count is None and full_comment is None:
                rows_count, full_comment = 0, ''

            self._execute(src_hook, 'dest_hook', context)
            context['task_instance'].xcom_push(key='transferSuccess', value=True)

            q_rows = context['task_instance'].xcom_pull(key='q_rows')
            rows_count = q_rows + rows_count
            full_comment = full_comment + f': {q_rows}, '
            context['task_instance'].xcom_push(key='full_comment', value=full_comment)
            context['task_instance'].xcom_push(key='full_qr', value=rows_count)
        except Exception as exc:
            context['task_instance'].xcom_push(key='transferSuccess', value=False)
            raise exc

    def give_xlsx_output(self, src_conn, dir_file, context):
        cursor, q_rows, target_rows = self.get_rows(context, src_conn)
        f = None
        decrypt_col_num = []
        if self.decrypt_col is not None:
            key = Variable.get("fernet_secret_key_asup")
            f = Fernet(key)
            for colno, heading in enumerate(cursor.description, start=0):
                if heading.name.upper() in self.decrypt_col:
                    decrypt_col_num.append(colno+1)

        logging.info("decrypt_col_num: %s rows", decrypt_col_num)

        wb = openpyxl.Workbook()
        wb.encoding = self.encoding
        ws = wb.active
        ws.title = 'Лист1'
        for colno, heading in enumerate(self.headings, start = 1):
            ws.cell(row = 1, column = colno).value = heading
        for rowno, row in enumerate(target_rows, start = 2):
            for colno, cell_value in enumerate(row, start = 1):
                val = None
                if colno in decrypt_col_num:
                    val = f.decrypt(cell_value.encode('utf-8')).decode('utf-8')
                else:
                    val = cell_value

                if self.check_date(val):
                    if val == '9999-12-31':
                        val = ''
                    else:
                        val = datetime.datetime.strptime(str(val), '%Y-%m-%d').strftime('%Y%m%d')

                ws.cell(row=rowno, column=colno).value = val
        with open(dir_file, mode=self.modewr) as tfile:
            wb.save(tfile)
        cursor.close()
        return q_rows


    def get_all_sheet_name(self, context, src_conn):
        cursor = src_conn.cursor()
        logging.info('execute sql: %s', self.sheet_name_sql)
        cursor.execute(query=self.render_template(self.sheet_name_sql, context=context), vars=self.parameters)
        sheetName = cursor.fetchall()
        logging.info('SheetName ALL: %s', sheetName)
        cursor.close()
        return sheetName

    def give_xls_output(self, src_conn, dir_file, context):
        cursor, q_rows, target_rows = self.get_rows(context, src_conn)

        wb = xlwt.Workbook(encoding=self.encoding if self.encoding else 'cp1251')
        style = xlwt.XFStyle()
        style.num_format_str = 'dd.mm.yyyy'

        if self.sheet_name_sql is not None:
            self.SheetName = self.get_all_sheet_name(context, src_conn)
            for _, row in enumerate(self.SheetName, start=0):
                for _, sheetName in enumerate(row, start=0):
                    ws = wb.add_sheet(sheetName, cell_overwrite_ok=True)
                    filter_rows = filter(lambda c: c[0] == sheetName, target_rows)
                    self.write_rows(style, filter_rows, ws)
        else:
            ws = wb.add_sheet(self.SheetName, cell_overwrite_ok=True)
            self.write_rows(style, target_rows, ws)

        with open(dir_file, mode=self.modewr) as tfile:
            wb.save(tfile)
        cursor.close()
        return q_rows

    def write_rows(self, style, target_rows, ws):
        for colno, heading in enumerate(self.headings, start=0):
            ws.write(r=0, c=colno, label=heading)
        for rowno, row in enumerate(target_rows, start=1):
            for colno, cell_value in enumerate(row, start=0):
                if isinstance(cell_value, datetime.date):
                    ws.write(r=rowno, c=colno, label=cell_value, style=style)
                else:
                    ws.write(r=rowno, c=colno, label=cell_value)

    def get_rows(self, context, src_conn):
        cursor = src_conn.cursor()
        cursor.execute(query=self.render_template(self.sql, context=context), vars=self.parameters)
        target_rows = cursor.fetchall()
        q_rows = len(target_rows)
        return cursor, q_rows, target_rows

    def check_date(self, date_text, date_format='%Y-%m-%d'):
        try:
            datetime.datetime.strptime(str(date_text), date_format)
            return True
        except ValueError:
            return False

    def give_excel_xml_output(self, src_conn, dir_file, context):
        cursor, q_rows, target_rows = self.get_rows(context, src_conn)
        f = None
        decrypt_col_num = []
        if self.decrypt_col is not None:
            logging.info("Decrypt_col: %s ", self.decrypt_col)
            key = Variable.get("fernet_secret_key_asup")
            f = Fernet(key)
            for colno, heading in enumerate(self.headings, start=0):
                if heading in self.decrypt_col:
                    decrypt_col_num.append(colno)

        logging.info("Fernet_KEY: %s ", f)
        tree = lxml.parse(self.template_excel_xml, lxml.XMLParser(remove_blank_text=True))
        root = tree.getroot()
        ns = root.nsmap
        ns.pop(None)

        ws = root.find('ss:Worksheet', ns)
        table = ws.find('ss:Table', ns)
        ss = ns.get('ss')

        for _, row in enumerate(target_rows, start=1):
            new_row = lxml.fromstring('''<Row xmlns:ss="%s" ss:AutoFitHeight="0"/>''' % ss)
            for colno, cell_value in enumerate(row, start=0):
                if cell_value is None:
                    val = ''
                elif colno in decrypt_col_num:
                    val = f.decrypt(cell_value.encode('utf-8')).decode('utf-8')
                else:
                    val = cell_value

                if self.check_date(val):
                    val = datetime.datetime.strptime(str(val), '%Y-%m-%d').strftime('%Y%m%d')

                new_cell = lxml.fromstring(
                    '''<Cell xmlns:ss="%s" ss:StyleID="s67"><Data ss:Type="String">%s</Data></Cell>''' % (ss, val))
                new_row.append(new_cell)
            table.append(new_row)
        table.set('{%s}ExpandedRowCount' % ss, str(q_rows+1))

        with open(dir_file, mode=self.modewr) as tfile:
            tree.write(tfile, pretty_print=True, encoding=self.encoding, xml_declaration=True)
        return q_rows

    # def give_excel_xml_output(self, src_conn, dir_file, context):
    #     cursor, f, q_rows, target_rows = self.method_name(context, src_conn)
    #
    #     decrypt_col_num = []
    #     for colno, heading in enumerate(self.headings, start=0):
    #         if heading in self.decrypt_col:
    #             decrypt_col_num.append(colno)
    #
    #     tree = lxml.parse(self.template_excel_xml, lxml.XMLParser(remove_blank_text=True))
    #     root = tree.getroot()
    #     ns = root.nsmap
    #     ns.pop(None)
    #
    #     ws = root.find('ss:Worksheet', ns)
    #     table = ws.find('ss:Table', ns)
    #     ss = ns.get('ss')
    #
    #     for _, row in enumerate(target_rows, start=1):
    #         new_row = lxml.fromstring('''<Row xmlns:ss="%s" ss:AutoFitHeight="0"/>''' % ss)
    #         for colno, cell_value in enumerate(row, start=0):
    #             val = cell_value if colno not in decrypt_col_num else f.decrypt(cell_value.encode('utf-8')).decode('utf-8')
    #             logging.info("Data cell_value %s ", val)
    #             new_cell = lxml.fromstring(
    #                 '''<Cell xmlns:ss="%s" ss:StyleID="s67"><Data ss:Type="String">%s</Data></Cell>''' % (ss, val))
    #             new_row.append(new_cell)
    #         table.append(new_row)
    #     table.set('{%s}ExpandedRowCount' % ss, str(q_rows+1))
    #
    #     with open(dir_file, mode=self.modewr) as tfile:
    #         tree.write(tfile, pretty_print=True, encoding=self.encoding, xml_declaration=True)
    #     return q_rows

    def in_development(*args):
        logging.info("This function is currently in development, stay tuned :)")
