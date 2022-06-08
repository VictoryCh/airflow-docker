import logging
import datetime

from airflow.providers.samba.hooks.samba import SambaHook
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.models import BaseOperator
from airflow.utils.decorators import apply_defaults
import openpyxl
import xlwt
import pandas as pd
# from lxml import etree as lxml
import xml.etree.ElementTree as ET


class Postgres2SMB(BaseOperator):
    template_fields = ('sql',)
    template_ext = ('.sql',)  # file format
    ui_color = '#e08c8c'
    dest_cols = []

    @apply_defaults
    def __init__(self, sql, postgres_conn_id=None, parameters=None, autocommit=False, rows_chunk=5000,
                 samba_conn_id=None,
                 headings=None, share=None, dir_samba=None, file_name='test', file_format='csv', delimiter=';',
                 encoding='utf-8', typ=None,
                 modewr='wb', SheetName='Лист1', template_xml=None, *args, **kwargs):
        super(Postgres2SMB, self).__init__(*args, **kwargs)
        if parameters is None:
            parameters = {}
        self.sql = sql  # sql query on source system (here, if necessary, transformations are performed)
        self.postgres_conn_id = postgres_conn_id
        self.parameters = parameters
        self.autocommit = autocommit
        self.rows_chunk = rows_chunk
        self.samba_conn_id = samba_conn_id
        self.headings = headings
        self.share = share
        self.dir_samba = dir_samba
        self.file_name = file_name
        self.file_format = file_format
        self.delimiter = delimiter
        self.encoding = encoding
        self.typ = typ
        self.modewr = modewr
        self.SheetName = SheetName
        self.template_xml = template_xml
        self.outputs = {'xml': self.get_xml_output,  # self.give_xml_output,
                        'csv': self.in_development,  # self.give_csv_output,
                        'xls': self.give_xls_output,
                        'xlsx': self.in_development  # self.give_xlsx_output
                        }

    def _execute(self, src_hook, dest_hook, context):
        with src_hook.get_conn() as src_conn:
            logging.info("Source id connection: %s", self.postgres_conn_id)
            logging.info("Executing sql = %s", self.sql)
            dirr = '' if not self.dir_samba else self.dir_samba + '/'
            dir_file = f'{dirr}{self.file_name}.{self.file_format}'
            q_rows = self.outputs[self.file_format](src_conn=src_conn, dir_file=dir_file, context=context)
            # q_rows = self.give_xls_output(src_conn, dir_file)

            logging.info("Finished data transfer")
            context['task_instance'].xcom_push(key='q_rows', value=q_rows)

    def execute(self, context):
        dest_hook = 'SambaHook(samba_conn_id=self.samba_conn_id, share=self.share)'
        src_hook = PostgresHook(postgres_conn_id=self.postgres_conn_id)
        try:
            rows_count = context['task_instance'].xcom_pull(key='full_qr')
            full_comment = context['task_instance'].xcom_pull(key='full_comment')
            if rows_count is None and full_comment is None:
                rows_count, full_comment = 0, ''

            self._execute(src_hook, 'dest_hook', context)
            context['task_instance'].xcom_push(key='transferSuccess', value=True)

            q_rows = context['task_instance'].xcom_pull(key='q_rows')
            rows_count = q_rows + rows_count
            full_comment = full_comment + f': {q_rows}, '
            context['task_instance'].xcom_push(key='full_comment', value=full_comment)
            context['task_instance'].xcom_push(key='full_qr', value=rows_count)
        except Exception as exc:
            context['task_instance'].xcom_push(key='transferSuccess', value=False)
            raise exc

    def give_csv_output(self, src_conn, dir_file, context):
        df = pd.read_sql_query(self.sql, con=src_conn)
        q_rows = len(df.index)
        logging.info("Data transfer: %s rows", q_rows)
        with open(dir_file, mode=self.modewr) as tfile:
            df.to_csv(tfile, sep=self.delimiter, header=self.headings, encoding=self.encoding)
        return q_rows

    def give_xlsx_output(self, src_conn, dir_file, context):
        cursor = src_conn.cursor()
        cursor.execute(query=self.render_template(self.sql, context=context), vars=self.parameters)
        target_rows = cursor.fetchall()
        q_rows = len(target_rows)
        logging.info("Data transfer: %s rows", q_rows)
        wb = openpyxl.Workbook(encoding=self.encoding)
        ws = wb.active
        ws.title = 'Лист1'
        for colno, heading in enumerate(self.headings, start=1):
            ws.cell(row=1, column=colno).value = heading
        for rowno, row in enumerate(target_rows, start=2):
            for colno, cell_value in enumerate(row, start=1):
                ws.cell(row=rowno, column=colno).value = cell_value
        with open(dir_file, mode=self.modewr) as tfile:
            wb.save(tfile)
        cursor.close()
        return q_rows

    def give_xls_output(self, src_conn, dir_file, context):
        cursor = src_conn.cursor()
        cursor.execute(query=self.render_template(self.sql, context=context), vars=self.parameters)
        target_rows = cursor.fetchall()
        q_rows = len(target_rows)
        logging.info("Data transfer: %s rows", q_rows)
        wb = xlwt.Workbook(encoding=self.encoding if self.encoding else 'cp1251')
        ws = wb.add_sheet(self.SheetName, cell_overwrite_ok=True)
        style = xlwt.XFStyle()
        style.num_format_str = 'dd.mm.yyyy'
        for colno, heading in enumerate(self.headings, start=0):
            ws.write(r=0, c=colno, label=heading)
        for rowno, row in enumerate(target_rows, start=1):
            for colno, cell_value in enumerate(row, start=0):
                if isinstance(cell_value, datetime.date):
                    ws.write(r=rowno, c=colno, label=cell_value, style=style)
                else:
                    ws.write(r=rowno, c=colno, label=cell_value)
        with open(dir_file, mode=self.modewr) as tfile:
            wb.save(tfile)
        cursor.close()
        return q_rows

    def get_xml_output(self, src_conn, dir_file, context):
        if self.template_xml is None:
            return self.give_xml_output(src_conn, dir_file, context)
        else:
            return self.give_excel_xml_output(src_conn, dir_file, context)


    def give_xml_output(self, src_conn, dir_file, context):
        root = ET.Element("root")
        cursor = src_conn.cursor()
        cursor.execute(query=self.render_template(self.sql, context=context), vars=self.parameters)
        target_rows = cursor.fetchall()
        q_rows = len(target_rows)

        for heading in self.headings:
            ET.SubElement(root, "header", value=heading)
        for _, row in enumerate(target_rows, start=1):
            for _, cell_value in enumerate(row, start=0):
                ET.SubElement(root, "cell", value=cell_value)
        tree = ET.ElementTree(root)
        with open(dir_file, mode=self.modewr) as tfile:
            tree.write(tfile)
        return q_rows

    def give_excel_xml_output(self, src_conn, dir_file, context):
        cursor = src_conn.cursor()
        cursor.execute(query=self.render_template(self.sql, context=context), vars=self.parameters)
        target_rows = cursor.fetchall()
        q_rows = len(target_rows)

        # tree = ET.parse(self.template_xml, ET.XMLParser(remove_blank_text=True))
        tree = ET.parse(self.template_xml)
        root = tree.getroot()
        # ns = root.nsmap
        # ns.pop(None)
        ns = {
            '': "urn:schemas-microsoft-com:office:spreadsheet",
            'o': "urn:schemas-microsoft-com:office:office",
            'x': "urn:schemas-microsoft-com:office:excel",
            'ss': "urn:schemas-microsoft-com:office:spreadsheet",
            'html': "http://www.w3.org/TR/REC-html40"
        }

        ws = root.find('ss:Worksheet', ns)
        table = ws.find('ss:Table', ns)
        ss = ns.get('ss')

        for _, row in enumerate(target_rows, start=1):
            new_row = ET.fromstring('''<Row xmlns:ss="%s" ss:AutoFitHeight="0"/>''' % ss)
            for _, cell_value in enumerate(row, start=0):
                new_cell = ET.fromstring(
                    '''<Cell xmlns:ss="%s" ss:StyleID="s67"><Data ss:Type="String">%s</Data></Cell>'''
                    % (ss, cell_value))
                new_row.append(new_cell)
            table.append(new_row)
        table.set('{%s}ExpandedRowCount' % ss, str(q_rows+1))

        for prefix, uri in ns.items():
            ET.register_namespace(prefix, uri)  # Ensure correct prefixes in output
            if prefix not in ("o", "x", "ss", "html", ""):  # Prevent duplicate ns declarations
                root.set("xmlns:" + prefix, uri)

        with open(dir_file, mode=self.modewr) as tfile:
            tree.write(tfile, encoding=self.encoding, xml_declaration=True)
        return q_rows

    def in_development(*args):
        logging.info("This function is currently in development, stay tuned :)")
